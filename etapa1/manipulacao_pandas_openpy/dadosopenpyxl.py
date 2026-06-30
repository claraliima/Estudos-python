import openpyxl 
import os
from openpyxl import Workbook # Importa a classe Worbook(tem como função criação de arquivos excel) da biblioteca openpyxl

PASTA = "arquivos"
CAMINHO_ARQUIVO = os.path.join(PASTA, "biblioteca_openpyxl.xlsx")
CAMPOS = ["isbn", "titulo", "autor", "genero", "ano_publicacao", "editora", "paginas", "status", "localizacao"]

def inicializar_diretorio():
    
    # Função para criar diretório caso não exista com nome da var PASTA
    if not os.path.exists(PASTA):
        os.makedirs(PASTA)

def salvar_no_arquivo(lista_dados):
    inicializar_diretorio()
    wb = Workbook() # 'wb' recebe um novo arquivo Excel vazio
    ws = wb.active # 'ws' recebe a aba ativa desse arquivo que está 'wb'
    ws.title = "Biblioteca" # Define o nome da aba

    ws.append(CAMPOS) # Adiciona uma linha na aba, onde cada coluna é um item da lista 'CAMPOS'

    # Escrever dados
    for livro in lista_dados:
        ws.append([livro.get(c, "") for c in CAMPOS]) #Cria lista para cada livro, essa lista se torna uma linha no arquivo representando cada livro. 'livro.get(c, "")' pega o valor da chave c no dicionário livro se livro não tiver essa chave, retorna "" (string vazia)
    
    wb.save(CAMINHO_ARQUIVO) # Salva arquivo que está em 'wb' 

def carregar_do_arquivo():
    if not os.path.exists(CAMINHO_ARQUIVO):
        return [] # Retorna uma lista vazia caso o 'CAMINHO_ARQUIVO' não exista, ou seja, não tem o que carregar do arquivo

    wb = openpyxl.load_workbook(CAMINHO_ARQUIVO) # Atribui a 'wb' o arquivo carregado(load_worbook) que tem como nome 'CAMINHO_ARQUIVO' e dá acesso as abas e células dentro dele
    
    ws = wb.active # 'ws' se torna a pasta ativa no arquivo 'wb'
    biblioteca_aux = []

    rows = list(ws.iter_rows(values_only=True)) # Lista que recebe todas as linhas da planilha como tupla. 'values_only=True' faz com que o retorno seja só o conteúdo das células
    
    if len(rows) <= 1: return [] # Só tem cabeçalho

    cabecalho = rows[0]
    for row in rows[1:]: # Pula a primeira linha(cabeçalho)
        
        livro = dict(zip(cabecalho, row)) # Zipa o cabeçalho com o valor da linha e converte esses pares em um dicionário
        
        biblioteca_aux.append(livro) #Adiciona o dicionário na lista
    
    return biblioteca_aux